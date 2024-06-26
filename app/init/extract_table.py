'''
https://pymupdf.readthedocs.io/en/latest/page.html#Page.find_tables
https://pypi.org/project/PyMuPDF/1.23.9/
PYDEVD_WARN_EVALUATION_TIMEOUT environment variable to a bigger value

https://www.docugami.com/pricing

https://github.com/langchain-ai/langchain/blob/master/cookbook/Semi_Structured_RAG.ipynb?ref=blog.langchain.dev 
https://python.langchain.com/docs/modules/data_connection/retrievers/multi_vector
https://levelup.gitconnected.com/a-guide-to-processing-tables-in-rag-pipelines-with-llamaindex-and-unstructuredio-3500c8f917a7


https://pymupdf.readthedocs.io/en/latest/recipes-text.html#how-to-extract-text-in-natural-reading-order

https://pymupdf.readthedocs.io/en/latest/recipes-text.html#how-to-analyze-font-characteristics


'''        

# take directory path as input and create entries in sqlalchemy db 
import os
import ast
import re
import io
import uuid
import hashlib
from dotenv import load_dotenv

# fitz is the PDF parser with table recognize capabilities 
import fitz
import pandas as pd
import time
from datetime import datetime, timedelta
from difflib import SequenceMatcher
from typing import List, Optional

# for drop box data access
import base64
import requests
import dropbox
import openpyxl as pxl 

# for more secure file handling
from werkzeug.utils import secure_filename

# for logging 
import logging
# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# LlmaIndex manages data and embeddings 
from llama_index import ServiceContext, VectorStoreIndex
# from llama_index.retrievers import VectorIndexRetriever
# from llama_index.query_engine import RetrieverQueryEngine
# from llama_index.postprocessor import SimilarityPostprocessor

# lower level functions for creating nodes 
from llama_index.schema import TextNode

# modules to create an http server with API
from flask import Flask, jsonify
from flask_sqlalchemy import SQLAlchemy
# https://docs.sqlalchemy.org/en/20/core/engines.html
from sqlalchemy import  Column, inspect
from sqlalchemy.exc import SQLAlchemyError

# queues for streaming updates during init 
from queue import Queue, Empty
import threading
# Global message queue
message_queue = Queue()

# get functions and variables created by __init__.py
from app import db, create_app
from app.models import ThreadComplete, User, Organization, Tables, TableData

# define some globals for dropbox resiliency 
MAX_RETRY_COUNT = 5
BACKOFF_FACTOR = 1.5
RETRY_STATUS_CODES = (503,)

# start of useful functions 

def string_to_list(string: str) -> List:
    """
    Converts a string representation of a list into an actual Python list object.

    Parameters:
    - string (str): A string that represents a list, e.g., "[1, 2, 3]".

    Returns:
    - List: The list object represented by the input string.

    Raises:
    - ValueError: If the input string does not represent a list or if there's an error
      in converting the string to a list.
    """
    
    try:
        result = ast.literal_eval(string)
    except (SyntaxError, ValueError) as e:
        # Catching specific exceptions related to literal_eval failures
        raise ValueError(f"Error converting string to list: {e}")
    
    if not isinstance(result, list):
        raise ValueError("The evaluated expression is not a list")

    return result

def find_indices_within_percentage(arr, percentage):
    """
    Finds the indices of the first occurrence where the difference between any three consecutive
    values in the list is less than a specified percentage of the average of these three values.
    
    Parameters:
    - arr: List of real numbers.
    - percentage: Specified percentage expressed as a decimal (e.g., 10% as 0.1).
    
    Returns:
    - A tuple of indices of the three values if such a case exists, otherwise None.
    """
    # Ensure the array is long enough
    if len(arr) < 3:
        return None

    for i in range(len(arr) - 2):
        # Extract three consecutive values
        a, b, c = abs(arr[i]), abs(arr[i + 1]), abs(arr[i + 2])
        
        # Calculate the average
        avg = (a + b + c) / 3
        
        # Calculate the maximum difference
        max_diff = max(abs(a - b), abs(a - c), abs(b - c))
        
        # Check if the maximum difference is less than the specified percentage of the average
        if max_diff < (avg * percentage):
            return (i, i + 1, i + 2)
    
    # Return None if no such trio is found
    return None

def find_longest_sorted_chain(tuples_list):
    """
    Finds the longest continuous chain of tuples where the second value of one tuple
    equals the first value of the next tuple. The input list is first sorted by the first
    element of each tuple as the primary key and the second element as the secondary key.

    Parameters:
    - tuples_list: A list of numeric tuples of length 2.

    Returns:
    - A list containing the longest continuous chain of tuples meeting the above criteria.

    Raises:
    - ValueError: If the input list contains items that are not tuples or tuples of incorrect length.
    """

    # Validate input list
    if not all(isinstance(item, tuple) and len(item) == 2 for item in tuples_list):
        raise ValueError("All items in the input list must be tuples of length 2.")
    
    # Sort tuples by the first value primarily and the second value secondarily
    sorted_tuples = sorted(tuples_list, key=lambda x: (x[0], x[1]))

    def find_chain(start_index):
        """Finds a continuous chain starting from the given index in the sorted list."""
        chain = [sorted_tuples[start_index]]
        last_value = sorted_tuples[start_index][1]

        for next_tuple in sorted_tuples[start_index + 1:]:
            if next_tuple[0] == last_value:
                chain.append(next_tuple)
                last_value = next_tuple[1]

        return chain

    longest_chain = []
    # Iterate over sorted tuples to find the longest chain
    for i in range(len(sorted_tuples)):
        current_chain = find_chain(i)
        if len(current_chain) > len(longest_chain):
            longest_chain = current_chain

    return longest_chain

def convert_epoch_to_datetime(epoch_seconds):
    """
    Convert epoch seconds to a human-readable datetime string.
    
    Parameters:
    - epoch_seconds: An integer or float representing the number of seconds since
                     the Unix epoch (January 1, 1970, 00:00:00 UTC).
    
    Returns:
    - A string representing the datetime in the format "Mon Jan 31 2024 7:34 PM".
    
    Raises:
    - TypeError: If the input is not an integer or float.
    - ValueError: If the epoch_seconds represent a date before the year 1900
                  or if the input is out of the valid range for a datetime.
    """
    # Validate input type
    if not isinstance(epoch_seconds, (int, float)):
        raise TypeError("epoch_seconds must be an integer or float")
    
    try:
        # Convert epoch time to a datetime object
        dt = datetime.fromtimestamp(epoch_seconds)
    except OverflowError as e:
        # Handle overflow error if epoch_seconds is out of range
        raise ValueError("epoch_seconds is out of valid range for a datetime") from e
    except OSError as e:
        # Handle OS errors, which might occur for extremely large or invalid values
        raise ValueError("Invalid epoch_seconds value") from e
    
    # Safely format the datetime object to "Mon Jan 31 2024 7:34 PM"
    formatted_time = dt.strftime('%a %b %d %Y %I:%M %p')
    
    return formatted_time

def convert_datetime_str_to_epoch(datetime_str):
    """
    Converts a datetime string in the format "Mon Jan 31 2024 7:34 PM" to epoch seconds.

    Parameters:
    - datetime_str: A string representing the datetime in the specific format
                    "%a %b %d %Y %I:%M %p", e.g., "Tue Mar 23 2021 04:50 PM".

    Returns:
    - An integer representing the number of seconds since the Unix epoch (January 1, 1970).

    Raises:
    - ValueError: If the datetime_str does not match the expected format or represents
                  a date that cannot be converted to epoch seconds.
    - TypeError: If the input is not a string.
    """
    # Validate input type
    if not isinstance(datetime_str, str):
        raise TypeError("datetime_str must be a string")

    try:
        # Attempt to parse the datetime string according to the specified format
        dt = datetime.strptime(datetime_str, '%a %b %d %Y %I:%M %p')
    except ValueError as e:
        # Handle cases where the string does not match the expected format
        raise ValueError("datetime_str does not match the expected format or is invalid") from e

    # Convert the datetime object to epoch seconds and return
    epoch_seconds = int(dt.timestamp())
    return epoch_seconds

def pdf_timestamp_to_readable(pdf_timestamp):
    """
    Converts a PDF-specific timestamp string to a human-readable datetime string.
    
    Parameters:
    - pdf_timestamp: A string in the PDF-specific timestamp format "D:YYYYMMDDhhmmss+hh'mm'" 
                     or "D:YYYYMMDDhhmmss-hh'mm'", where the timezone is optional.
    
    Returns:
    - A string representing the datetime in the format "Mon Jan 31 2024 7:34 PM".
    
    Raises:
    - ValueError: If the pdf_timestamp does not conform to the expected format or
                  contains invalid date/time values.
    - TypeError: If the input is not a string.
    """
    # Validate input type
    if not isinstance(pdf_timestamp, str):
        raise TypeError("pdf_timestamp must be a string")
    
    # Initial format check
    if not pdf_timestamp.startswith('D:') or len(pdf_timestamp) < 16:
        raise ValueError("pdf_timestamp does not conform to the expected format")

    # Extract timestamp and timezone
    ts = pdf_timestamp[2:16]  # YYYYMMDDhhmmss
    tz = pdf_timestamp[16:]   # Time zone information

    try:
        # Convert the timestamp to a datetime object
        dt = datetime.strptime(ts, '%Y%m%d%H%M%S')
    except ValueError as e:
        raise ValueError("Invalid timestamp format or value") from e

    if tz:
        try:
            # Parse timezone information
            sign = tz[0]
            tz_hours = int(tz[1:3])
            tz_minutes = int(tz[4:6])
            delta = timedelta(hours=tz_hours, minutes=tz_minutes)

            # Adjust datetime object according to the timezone
            if sign == '+':
                dt -= delta
            elif sign == '-':
                dt += delta
        except (ValueError, IndexError) as e:
            raise ValueError("Invalid timezone format or value") from e

    # Format and return the datetime object
    formatted_time = dt.strftime('%a %b %d %Y %I:%M %p')
    return dt, formatted_time

def string_to_datetime(date_string):
    """
    Converts a string representing a datetime in the format "Mon Jan 31 2024 7:34 PM"
    to a datetime object.
    
    Parameters:
    - date_string: A string representing the datetime in the format
                   "Mon Jan 31 2024 7:34 PM".
    
    Returns:
    - A datetime object representing the specified date and time.
    
    Raises:
    - ValueError: If the date_string does not match the expected format or
                  represents an invalid date/time.
    - TypeError: If the input is not a string.
    """
    # Validate input type
    if not isinstance(date_string, str):
        raise TypeError("date_string must be a string")

    try:
        # Convert the string to a datetime object
        return datetime.strptime(date_string, '%a %b %d %Y %I:%M %p')
    except ValueError as e:
        raise ValueError("date_string does not match the expected format or is invalid") from e
    
def get_last_modified_datetime(fullfilename):
    modt = os.stat(fullfilename).st_mtime
    date_string = convert_epoch_to_datetime(modt)
    return date_string

def allowed_file(filename):
    '''
    CHeck filename against allowed extensions 
    '''
    ALLOWED_EXTENSIONS = {'txt', 'pdf', 'xls', 'xlsx'}

    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def most_frequent_integer(int_list: List[int]) -> Optional[int]:
    """
    Finds the most frequently occurring integer in a list.
    
    Parameters:
    - int_list (List[int]): A list of integers to analyze.
    
    Returns:
    - Optional[int]: The integer that appears most frequently in the list. If there are multiple integers
      with the same highest frequency, one of them will be returned. Returns None if the list is empty
      or contains no integers.
    
    Raises:
    - ValueError: If the input list contains non-integer elements.
    - ValueError: If the input is not a list.
    """
    
    if not isinstance(int_list, list):
        raise ValueError("Input must be a list.")
    
    if len(int_list) == 0:
        return None  # Or raise an exception, depending on desired behavior
    
    if not all(isinstance(x, int) for x in int_list):
        raise ValueError("The list must contain only integers.")

    frequency = {}
    for num in int_list:
        frequency[num] = frequency.get(num, 0) + 1

    entries = sorted(frequency.items(), key=lambda x: (x[1], x[0]), reverse=True)

    return entries[0][0] # largest key of most frequently occurring col count
    # return most_frequent

def most_frequent_num(float_list: List[float]) -> Optional[int]:
    """
    Finds the most frequently occurring integer in a list.
    
    Parameters:
    - float_list (List[int]): A list of numbers to analyze.
    
    Returns:
    - Optional[int]: The number that appears most frequently in the list. If there are multiple numbers
      with the same highest frequency, one of them will be returned. Returns None if the list is empty
      or contains no numbers.
    
    Raises:
    - ValueError: If the input list contains non-float elements.
    - ValueError: If the input is not a list.
    """
    
    if not isinstance(float_list, list):
        raise ValueError("Input must be a list.")
    
    if len(float_list) == 0:
        return None  # Or raise an exception, depending on desired behavior
    
    if not all(isinstance(x, float) for x in float_list):
        raise ValueError("The list must contain only integers.")

    frequency = {}
    for num in float_list:
        frequency[num] = frequency.get(num, 0) + 1

    most_frequent = max(frequency, key=frequency.get, default=None)

    return most_frequent

def parse_fullfilename(full_filename:str = None):
    '''
    Parse a full file URI into components
    : returns a dictionary 
        'fullfilename': full_filename,
        'dirpath':dirpath, 
        'vendor':vendor,
        'filename':filename, 
        'filetoken': filetoken, # secure_filename(filename.split('.')[0])
        'filetype':filetype
        'worksheet': None # each sheet in an XL workbook is a unique table
        'dropboxdbx' : None (placeholder for dropbox dbx object)
        'dropboxid': None (placeholder for dropboxID)
        'dropboxurl': None (placeholder for dropbox url)
    '''
    try:
        # get the filename and directory path and file type
        dirpath = '/'.join(full_filename.split('/')[0:-1])    
        vendor = dirpath.split('/')[-1]
        filetype = secure_filename(full_filename).split('.')[-1]
        filename = full_filename.split('/')[-1]
        filetoken = secure_filename(filename.split('.')[0])

        return {'fullfilename': full_filename,'dirpath':dirpath,'filename':filename, 
                'filetoken':filetoken, 'filetype':filetype, 'vendor':vendor, 'worksheet':None,
                'dropboxdbx':None, 'dropboxid':None, 'dropboxurl':None  }
    except:
        return {'fullfilename':None,'dirpath':None, 'filename':None, 
                'filetoken':None, 'filetype':None, 'vendor':None, 
                'dropboxdbx':None, 'dropboxid':None, 'dropboxurl':None   }

def get_dropbox_accesstoken_from_refreshtoken(REFRESH_TOKEN, APP_KEY, APP_SECRET):
    '''
    ### Get the auth token from a valid refresh token, Key and Secret
    returns an authorization token for dropbox, error otherwise 

    https://www.dropboxforum.com/t5/Dropbox-API-Support-Feedback/Get-refresh-token-from-access-token/m-p/596755/highlight/false#M27728
    https://www.dropbox.com/developers/documentation/http/documentation#authorization 
    https://www.dropbox.com/oauth2/authorize?client_id=j27yo01f5f8w304&response_type=code&token_access_type=offline
    
    i have to get the access code dynamically from the refresh token
    https://developers.dropbox.com/oidc-guide 

    This gets a refresh token if we pass a valid AUTHORIZATIONCODEHERE
    We only need to do this once to get a persistent refresh token  
    curl https://api.dropbox.com/oauth2/token \
        -d code=AUTHORIZATIONCODEHERE \
        -d grant_type=authorization_code \
        -u APPKEYHERE:APPSECRETHERE​

    This gets the authcode from the refresh token
    curl https://api.dropbox.com/oauth2/token \
        -d refresh_token=REFRESHTOKENHERE \
        -d grant_type=refresh_token \
        -d client_id=APPKEYHERE \
        -d client_secret=APPSECRETHERE
    '''
    try:

        if not all([APP_KEY, APP_SECRET, REFRESH_TOKEN]):
            logging.error("Missing required Dropbox configuration.")
            return "Error: Missing configuration."

        BASIC_AUTH = base64.b64encode(f'{APP_KEY}:{APP_SECRET}'.encode()).decode()

        data = {
            "refresh_token": REFRESH_TOKEN,
            "grant_type": "refresh_token",
            "client_id": APP_KEY,
            "client_secret": APP_SECRET
        }

        response = requests.post('https://api.dropboxapi.com/oauth2/token', data=data)

        if response.status_code == 200:
            return response.json().get("access_token")
        else:
            logging.error(f"Dropbox API Error: {response.status_code} - {response.text}")
            return "Error: Failed to retrieve access token."

    except requests.RequestException as e:
        logging.error(f"Request Exception: {e}")
        return "Error: Network error occurred."

    except Exception as e:
        logging.error(f"General Exception: {e}")
        return "Error: An unexpected error occurred."

def is_file_dropbox(dropboxMeta):
    '''
    Check a drop box object id and test if it is a file
    Returns True if object is file, False otherwise
    '''
    return isinstance(dropboxMeta,dropbox.files.FileMetadata)

def get_dropbox_filenames(dbx, startDir:str):
    """
    Walks down all subdirectories from a top-level directory and collects filenames.

    :param startDir: The top-level directory to start the traversal from.
    :return: A Dict of filenames found in all subdirectories.
    """
    filesFound = {} # key = filepath, value = dbx meta entity

    # Recursive helper function
    def walk_dir_recursive(dbx, current_dir:str):
        try:
            entries = dbx.files_list_folder(current_dir).entries
            for entry in entries:
                if is_file_dropbox(entry):
                    logging.info(f"In {f'/{current_dir}'} found File: {entry.name}")
                    filesFound[entry.path_lower] = entry
                else: 
                    logging.info(f"Recursing into {f'/{entry.path_lower}'} ")
                    walk_dir_recursive(dbx, entry.path_lower)  # Recursion step

        except PermissionError:
            logging.error(f"Permission denied: {current_dir}")
            logging.error(f"Error accessing Dropbox: {e}")
        except FileNotFoundError:
            logging.error(f"File not found: {current_dir}")
            logging.error(f"Error accessing Dropbox: {e}")
        except OSError as e:
            logging.error(f"OS error: {e}")
            logging.error(f"Error accessing Dropbox: {e}")

        return(filesFound)
    
    walk_dir_recursive(dbx, startDir)

    return filesFound

def create_shared_link_with_retries(dbx, file_path):
    for attempt in range(MAX_RETRY_COUNT):
        try:
            shared_link_meta = dbx.sharing_create_shared_link(file_path)
            logging.info(f"Shared Link: {shared_link_meta.url}")
            return shared_link_meta.url  # Ensure the URL is returned on success
        except dropbox.exceptions.ApiError as api_err:
            handle_api_error(api_err, attempt)

def handle_api_error(api_err, attempt, file_path):
    if api_err.error.is_path() and api_err.error.get_path().is_not_found():
        raise Exception(f'File {file_path} does not exist.')
    elif api_err.status_code in RETRY_STATUS_CODES:
        wait_seconds = (2 ** attempt) * BACKOFF_FACTOR
        logging.error(f"HttpError status_code={api_err.status_code}: Waiting for {wait_seconds} seconds...")
        time.sleep(wait_seconds)
    else:
        raise api_err

def get_filenames_in_directory(top_dir):
    """
    Walks down all subdirectories from a top-level directory and collects filenames.

    :param top_dir: The top-level directory to start the traversal from.
    :return: A list of filenames found in all subdirectories.
    """
    filesFound = {}

    # Recursive helper function
    def walk_dir_recursive(current_dir):
        try:
            for entry in os.scandir(current_dir):
                if entry.is_file() and entry.name[0:2] != '~$':
                    filesFound[entry.path] = entry.path
                    #filenames.append(entry.path)
                elif entry.is_dir():
                    walk_dir_recursive(entry.path)  # Recursion step
        except PermissionError:
            logging.error(f"Permission denied: {current_dir}")
            logging.error(f"Error accessing Dropbox: {e}")
        except FileNotFoundError:
            logging.error(f"File not found: {current_dir}")
            logging.error(f"Error accessing Dropbox: {e}")
        except OSError as e:
            logging.error(f"OS error: {e}")
            logging.error(f"Error accessing Dropbox: {e}")

    walk_dir_recursive(top_dir)

    return filesFound

def get_firstpage_tables_as_list(doc=None):
    '''
    Get the first page of a fitz doc as a list of lists 
    : tbl_out is the list of lists
    : col_count is the number of columns (max) 
    :table_strategy(lines or text )is the fitz table parsing strategy used'''

    if not(doc):
        raise ValueError("fitz document not set - get_firstpage_tables_as_list")

    # focus on the first page 
    table_strategy = 'lines'
    tbls = doc[0].find_tables(vertical_strategy='lines', horizontal_strategy='lines')
    if len(tbls.tables) ==0: # did not find tables by grid, try spaces 
    # if True: # did not find tables by grid, try spaces 
        tbls = doc[0].find_tables(vertical_strategy='text', horizontal_strategy='text')
        table_strategy = 'text'

    # merge the tables 
    tbl_out = []
    col_counts = []

    for tbl in tbls.tables:
        tbl_out.extend(tbl.extract())
        col_counts.append(tbl.col_count)

    return tbl_out, table_strategy, tbls

def find_matching_index_ignore_case(list1, list2):
    """
    Finds the index of the first list in list2 where any element of list1 matches any element of the list at that index,
    ignoring case. This function also handles NaN values which might be present in the lists derived from pandas.

    Parameters:
    - list1 (list of str): List of strings to search for.
    - list2 (list of list of str): List of lists where each entry is a list of strings.

    Returns:
    - int: The index of the first matching list in list2, or None if no match is found.
    """
    # Convert all items in list1 to lowercase for case-insensitive comparison
    list1_lower = [str(item).lower() for item in list1 if item is not pd.NA and pd.notna(item)]
    for index, sublist in enumerate(list2):
        # Convert all items in the current sublist to lowercase, skipping NaN values
        sublist_lower = [str(item).lower() for item in sublist if item is not pd.NA and pd.notna(item)]
        if any(item in sublist_lower for item in list1_lower):
            return index
    return None

def cell_diff(cells: list) -> list:
    """
    Calculate the element-wise differences between consecutive vectors in a list.
    
    Parameters:
    - cells: A list of lists (vectors) where each inner list contains numerical values.
    
    Returns:
    - A list of lists containing the rounded element-wise differences between consecutive vectors.
    
    Raises:
    - ValueError: If 'cells' contains less than two vectors or if any vector contains non-numeric values.
    - TypeError: If 'cells' is not a list of lists.
    """
    
    # Check if 'cells' is a list of lists
    if not all(isinstance(cell, tuple) for cell in cells):
        raise TypeError("All elements in 'cells' must be lists.")
    
    # Check if 'cells' has at least two vectors
    if len(cells) < 2:
        raise ValueError("The 'cells' list must contain at least two vectors to compute differences.")
    
    # Ensure all elements in each vector are numeric
    for vec in cells:
        if not all(isinstance(num, (int, float)) for num in vec):
            raise ValueError("All elements in each vector must be numeric.")

    def vec_diff(vec1, vec2):
        """Calculate and return the rounded element-wise difference between two vectors."""
        return [round(v1 - v2, 2) for v1, v2 in zip(vec1, vec2)]
    
    # Compute the differences between consecutive vectors
    c_diff = [vec_diff(cells[n-1], cells[n]) for n in range(1, len(cells))]

    return c_diff

def get_header_match_table_cells(fitz_page=None, fitz_table=None):
    """
    Extracts header text matches for each table cell from a provided page and table object.

    Parameters:
    - fitz_page: A page object from PyMuPDF (fitz) library, containing the table.
    - fitz_table: A table object, typically extracted from a page, which contains cells.

    Returns:
    - A tuple containing the table header grid and the start index of the table body,
      or None if an error occurs or input parameters are invalid.
    """
    if fitz_table is None or fitz_page is None:
        logging.error("Invalid input: 'fitz_table' or 'fitz_page' is None.")
        return None

    try:
        hdr = fitz_table.header
        hdr_0 = hdr.cells[0]

        first_row = 0 if fitz_table.header.external else 1

        cell_d = cell_diff(fitz_table.cells)
        tbl_start = find_indices_within_percentage([x[1] for x in cell_d], 0.02)
        grid_y = round(sum([abs(x[1]) for x in cell_d[tbl_start[0]:tbl_start[2]]]) / 2, 1)

        cell_x = set([(x[0], x[2]) for x in fitz_table.cells])
        try:
            grid_x = find_longest_sorted_chain(cell_x)
        except ValueError as e:
            logging.error(f"find_longest_sorted_chain error {e}")

        tbl_hdr_rows = []
        for row in range(5):  # Iterate over an arbitrary number of rows above grid start
            y_row = (fitz_table.cells[tbl_start[0]][1] - row * grid_y,
                     fitz_table.cells[tbl_start[0]][1] - (row + 1) * grid_y)
            tbl_cells = [(x[0], y_row[1], x[1], y_row[0]) for x in grid_x]
            tbl_hdr = [fitz_page.get_textbox(c).strip() for c in tbl_cells]
            tbl_hdr_rows.append(tbl_hdr)

        tbl_hdr_grid = tbl_hdr_rows[::-1]
        return tbl_hdr_grid, tbl_start[0]

    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        return None

def get_bestguess_table_header(table_as_list = None):
    '''
    Uses embedding and LLM to guess the best header row
    : returns a dictionary
        'header_rownum' : header_rownum,
        'header_node_text' : header_node_text,
        'header_raw' : header_raw,
        'p_key' : p_key, a variable with the name of the sqlalchemy table primary key 
        'header_guess' : header_guess
        'plantname_in_header':plantname_in_header
    '''

    if not isinstance(table_as_list, list) or not all(isinstance(row, list) for row in table_as_list):
        logging.error("Invalid input: table_as_list must be a list of lists.")
        return None
    
    if len(table_as_list) == 0:
        logging.info("Empty table provided.")
        return None

    try:
        # get the best guess at the header row
        if len(table_as_list) > 0:

            # guess the most likely number of columns from the most frequent table column count 
            numcols = most_frequent_integer([len(x) for x in table_as_list])

            # find the first row populated with most frequent number of columns
            table_row_start_guess = 0
            for n,row in enumerate(table_as_list):
                if len([r for r in row if r ==r]) >= int(numcols*7/10):
                    table_row_start_guess = n
                    break

            if table_row_start_guess is None:
                logging.error("No suitable row start guess found.")
                return None

            # create text nodes with one node per row in tmp 
            table_nodes = [TextNode(text=f"'{t}'", id_=n) for n, t in enumerate(table_as_list) if isinstance(t, list)]
            if not table_nodes:
                logging.error("Failed to create text nodes from table.")
                return None
            table_index = VectorStoreIndex(table_nodes)

            # create the query engine with custom config to return just one node
            # https://docs.llamaindex.ai/en/stable/module_guides/deploying/query_engine/root.html#query-engine 
            query_engine = table_index.as_query_engine(
                similarity_top_k=1,
                vector_store_query_mode="default",
                alpha=None,
                doc_ids=None,
            )
            # request gpt's best guess at the row to use for table headings 
            multishotprompt=f"""Return a rows with as close to {numcols} table column headings as you can find 
            for this price sheet plants as a python list of length as close to {numcols} as feasible. 

            Examples of valid table rows with column headings:
            2 column: ['Description', 'A']
            5 column: ['Name', 'Latin Name', 'Price', 'Available Qty', 'Order Qty']
            8 column: [ "Product","SIZE1","SIZE2","PRICE", "AVL",  "COMMENTS", "ORDER", "Total"]
            7 column : ["Category", "WH", "Code", "Botantical Name", "size", "Price", "Available"]
            6 column: ['Order Qty', 'Part number', 'Description', 'UPC Code', 'WH $ 2024\nNET', 'COMMENTS']
            7 column: ['Code', 'Botanical', 'Size', 'Quantity', 'Price', '25 plus price', 'Order']
            Return an existing row. Do not make up rows.
            """
            
            try:
                response = query_engine.query(multishotprompt)
            except:
                logging.error(f"Error - gpt guess table column headings {response}")
                return None

            header_rownum = int(response.source_nodes[0].id_)
            header_node_text = response.source_nodes[0].text
            header_raw = table_as_list[header_rownum]
            header_gpt = string_to_list(response.response)

            # correct for special case failure of table header guess 
            # try the gpt suggestion first 
            if len(header_gpt) == numcols: 
                header_guess = ['TableName'] + header_gpt + ['Text']
                gpt_hdr_index = find_matching_index_ignore_case(header_gpt, table_as_list)
                if gpt_hdr_index is not None:
                    header_rownum=gpt_hdr_index
            elif len(header_raw) == numcols:
                header_guess = ['TableName'] + [str(x).replace(' ','').replace('\n','_').replace('(','_').replace(')','')
                        for x in header_raw] + ['Text']
            else:
                # this excludes header guesses with a different number of headings 
                header_guess = ['TableName'] + [f'col_{n}' for n in range(0,numcols)] + ['Text']
                header_rownum = table_row_start_guess # start of the frequently ocurring table width 

            best_guess={    
            'header_rownum' : header_rownum,
            'header_node_text' : table_nodes[header_rownum].text,
            'header_raw' : header_raw,
            'header_guess' : header_guess,
            'plantname_in_header': 'NotUsed'
            }
            
            return best_guess
        else:
            return None
    except :
        logging.error(f"Error - get_bestguess_table_header")
        return None
        pass

def get_bestguess_table_search_columns(file_table):
    # Use the table to guess which columns have plant names and descriptions
    # submit with first few rows to guess columns having plant names or botantical names 
    try:
        # Generate the desired dictionary with keys as column names
        col_dict = {
            column_name: ', '.join(file_table[column_name].head(5).astype(str)) 
            for column_name in file_table.columns if column_name not in
            ['id_hash', 'TableName']
        }

        # create text nodes with one node per dict entry
        dict_nodes = [TextNode(text=f"Dict Key: {str(k)} has values: {str(v)}", id_=k) 
                        for k,v in col_dict.items()]
        dict_index = VectorStoreIndex(dict_nodes)

        # create the query engine 
        query_engine = dict_index.as_query_engine(
            similarity_top_k=3,
            vector_store_query_mode="default", alpha=None, doc_ids=None )

        # get the gpt's best guess at the columns containing plant names and descriptions
        oneshotprompt=f"""You are an expert on garden center plants and their botantical names.
        Return the dictionary keys that contain descriptions of plants or botanical names. 
        Return only existing keys. Do not explain. Just return the key.
        """

        try:
            response = query_engine.query(oneshotprompt)
        except:
            logging.error(f"Error - gpt guess table search columns {response}")
            return None
        
        # df_search_cols = response.response.split(',')
        df_search_cols = list(response.metadata.keys())
        jnk = 0

        return df_search_cols

    except ValueError as e:
        print(f"Value Error best_guess_table_search_columns: {e}")
        return None
    
def get_file_table_pdf(file_data:dict = None):
    '''
    Input: Directory path and filename 
    Output: Dict
        dataframe 
        table_header = {'filename': filename,
                    'last_mod_datetime': last modified date-time as 'Mon Jan 31 2024 7:34 PM'
                    'numcols' : numcols,
                    'header_rownum':header_rownum,
                    'header_guess':header_guess, 
                    'header_raw':header_raw, 
                    'header_node_text':header_node_text}
        name of table's primary key 
    '''
    try:
        # check if the file exists 
        full_filename = file_data.get('fullfilename')

        # get the best guess at the header row
        if len(full_filename) > 0:
            dbx_file = None
            # branch based on the source of the files 
            if file_data.get('dropboxdbx') and file_data.get('dropboxid'):
                try:
                    dbx = file_data.get('dropboxdbx')
                    mfile = file_data.get('dropboxid')
                    # retrieve modified datetime and object to read file 
                    last_mod_datetime = mfile.client_modified
                    last_mod_datetime_string = last_mod_datetime.strftime('%a %b %d %Y %I:%M %p')
                    # download the dbx file object
                    md,dbxfile = dbx.files_download(mfile.path_lower)
                    # open a buffer to the dbx file 
                    dbx_file = io.BytesIO(dbxfile.content) 
                    # PDF: read the doc and get the first page tables
                    doc = fitz.open(stream=dbx_file.read(), filetype="pdf")
                except:
                    logging.error(f'Error retrieving dropbox file {full_filename} ')
                    return None, None
            else:
                # get file os last modified date time as string like 'Mon Jan 31 2024 7:34 PM'
                # last_modified = get_last_modified_datetime(full_filename) 
                # retrieve the metadate last modified date 
                tmp = doc.metadata['modDate']
                last_mod_datetime, last_mod_datetime_string =  pdf_timestamp_to_readable(tmp)
                del tmp
                # PDF: read the doc and get the first page tables
                doc = fitz.open(full_filename)

            # logging.info(f"get_file_table_pdf OS mod = {last_mod_datetime} PDF mod = {last_mod_datetime_string} ")

            # get the merged tables from first page 
            tables_aslist, tbl_strategy, tbls = get_firstpage_tables_as_list(doc)
            # guess the dominant column count
            numcols = most_frequent_integer([len(x) for x in tables_aslist])

            # special case for PDF header with different number of cells
            fitz_header_aslist = None
            for t in tbls.tables:
                # is our table the right width? If so, it is our main table
                if t.col_count == numcols:
                    fitz_header = t.header
                    # is the header the right width?
                    if len(fitz_header.cells) == t.col_count:
                        # is the header already in the table array?
                        if fitz_header.external:
                            fitz_header_aslist = fitz_header.names
                            tables_aslist = [fitz_header.names] + tables_aslist
                            break
                        else:
                            break
                    else:
                        # different cell counts for header and table 
                        # parse area above the table with the table cell boundaries 
                        fitz_header_aslist, tbl_start_row = get_header_match_table_cells(fitz_page = doc[0], 
                                                                          fitz_table = t)

                        # add the parsed headers to the table body 
                        tables_aslist = fitz_header_aslist + tables_aslist[tbl_start_row:]

            # use GPT to guess the header row - return dict 
            best_guess = get_bestguess_table_header(tables_aslist)

            col_len_test = len(best_guess.get('header_guess'))-2  # added Text, Tablename as columns in best_guess

            # merge header with the table from the first page 
            table_data = [ [file_data.get('filename')] +  row + [f"{row}"] 
                          for row in tables_aslist[best_guess.get('header_rownum')+1:] 
                          if len(row) == col_len_test]
            
            # add the tables from the rest of the pages using the same tbl_strategy 
            # for pagenum in range(1,doc.page_count):
            for pagenum in range(1,min(200,doc.page_count)):
                logging.info(f"Processing page {pagenum} of {doc.page_count}")

                tbls = doc[pagenum].find_tables(vertical_strategy=tbl_strategy, 
                                                horizontal_strategy=tbl_strategy)
                
                # exclude rows identical to the header row 
                tbl_page = [row for tbl in tbls.tables for row in tbl.extract() 
                            if row != best_guess.get('header_raw')]
                
                # add col of row as text and add filename and text string 
                table_data.extend([ [file_data.get('filename')] + row + [f"{row}"] for row in tbl_page 
                                   if len(row) == col_len_test ])
                
                jnk = 0 # for debugging

            # close the dropbox byteio 
            if dbx_file:
                dbx_file.close()
                del dbx_file

            # create a pandas dataframe 
            df = pd.DataFrame(table_data, columns = best_guess.get('header_guess'))
            # df.set_index('Table', inplace=True, drop=False)

            # save the file header info
            table_header = {'filename': file_data.get('filename'),
                            'last_mod_datetime':last_mod_datetime, 
                            'numcols' : len(df.columns),
                            'header_rownum':best_guess.get('header_rownum'),
                            'header_guess':best_guess.get('header_guess'), 
                            'header_raw':best_guess.get('header_raw'), 
                            'header_node_text':best_guess.get('header_node_text'),
                            'plantname_in_header':best_guess.get('plantname_in_header')}
                        
            return {full_filename:df}, {full_filename:table_header}
        else:
            logging.error('No table found with fitz parse by grid or text ')
            return None, None
    except:
        logging.error('get_file_table_pdf error ')
        return None, None
    
def get_file_table_xls(file_data = None):
    '''
    Input: Directory path and filename 
    Output: Dict
        dataframe 
        table_header = {'filename': filename,
                    'last_mod_datetime': last modified date-time as 'Mon Jan 31 2024 7:34 PM'
                    'numcols' : numcols,
                    'header_rownum':header_rownum,
                    'header_guess':header_guess, 
                    'header_raw':header_raw, 
                    'header_node_text':header_node_text}
        name of table's primary key 
    '''
    try:
        # check if the file exists 
        full_filename = file_data.get('fullfilename')
        if len(full_filename) > 0:
            # branch based on the source of the files 
            if file_data.get('dropboxdbx') and file_data.get('dropboxid'):
                try:
                    dbx = file_data.get('dropboxdbx')
                    mfile = file_data.get('dropboxid')
                    # retrieve modifed datetime and object to read file 
                    last_modified = mfile.client_modified
                    md,dbxfile = dbx.files_download(mfile.path_lower)
                    # open a buffer to the dbx file 
                    dbx_file = io.BytesIO(dbxfile.content) 
                    # load the workbook 
                    wb = pxl.load_workbook(dbx_file)
                    # dict with a key for each worksheet 
                    # wb_sheets = {ws: wb[ws] for ws in wb.sheetnames}
                    wb_sheets = {ws: [value for n,value in enumerate(wb[ws].iter_rows(values_only=True)) 
                                      if wb[ws].row_dimensions[n].hidden == False]  
                                 for n,ws in enumerate(wb.sheetnames)}

                    # 2024-06-05 - check for hidden rows - log results 
                    ws_rows_visible = {ws: [nn for nn,r in enumerate(wb[ws].rows) 
                                            if wb[ws].row_dimensions[nn].hidden == False] 
                             for n,ws in enumerate(wb.sheetnames) }
                    ws_rows_hidden = {ws: [nn for nn,r in enumerate(wb[ws].rows) 
                                           if wb[ws].row_dimensions[nn].hidden == True] 
                             for n,ws in enumerate(wb.sheetnames) }                    #  log results 
                    logging.info(f'Worksheets {[(k,len(v),"visible rows") for k,v in ws_rows_visible.items()]} ')
                    logging.info(f'Worksheets {[(k,len(v),"hidden rows") for k,v in ws_rows_hidden.items()]} ')

                    # df is dict - map each worksheet to a dataframe 
                    df_dict = {nm:pd.DataFrame(val) for nm,val in wb_sheets.items()}
                    dbx_file.close()

                except:
                    logging.error(f'Error retrieving dropbox file {full_filename} ')
                    logging.error(f"DBX {type(file_data.get('dropboxdbx'))}" )
                    logging.error(f"DBX fileid {type(file_data.get('dropboxid'))} '")
                    return None, None
            else:
                # 2024-06-06 - no support of multiple sheets - pd.read_excel only works for one sheet workbooks 
                # get the last modified date time as string like 'Mon Jan 31 2024 7:34 PM'
                last_modified = get_last_modified_datetime(full_filename)
                # read file as dataframe 
                df = pd.read_excel( full_filename, header=None)

            # create dict for headers - one for each sheet 
            table_header = {}
            # iterate over all excel sheets 
            for nm, df in df_dict.items():
                # remove columns with all null values 
                df.dropna(axis=1, how='all', inplace=True)
                # remove rows with all null values 
                df.dropna(axis=0, how='all', inplace=True)

            # get_bestguess_table_header => returns a dictionary
                best_guess = get_bestguess_table_header(table_as_list=df.head(50).values.tolist())
                if best_guess is None:
                    logging.error(f"Invalid input: {full_filename} raised error in get_bestguess_table_header.")

                # Convert row values to list as original text of row 
                row_text = df.apply(lambda row: str(list(row.values)), axis=1)
                # Add the filename and row as text to the DataFrame
                df['rowtext'] = row_text # add as last column 
                del row_text
                # Add the filename as first column to the DataFrame
                df.insert(0, 'filename', file_data.get('filename'))
                # update the column names 
                df.columns = best_guess.get('header_guess')

                # replace all null values with a string 
                df.fillna('empty', inplace=True)
                # drop all the rows above the header row
                df = df.iloc[best_guess.get('header_rownum')+1:]

                table_header[nm] = {'filename': file_data.get('filename'),
                                'last_mod_datetime':last_modified,
                                'numcols' : len(df.columns),
                                'header_rownum':best_guess.get('header_rownum'),
                                'header_guess':best_guess.get('header_guess'), 
                                'header_raw':best_guess.get('header_raw'), 
                                'header_node_text':best_guess.get('header_node_text'),
                                'plantname_in_header':best_guess.get('plantname_in_header')}
            
            return df_dict, table_header
        else:
            print('No table found with fitz parse by grid or text ')
            return None, None 
    except:
            pass    

def save_table_to_session(file_data, file_table, table_header, search_headers):
    """
    Updates or creates entries in the Tables and TableData session based on provided data.
    Does not commit objects -  db.session.commit() required 

    Parameters:
    - file_data: Dict keys: 'fullfilename', 'dirpath', 'vendor', 'filename', 'filetoken', 
                    'filetype', dropboxdbx, dropboxid, dropboxurl.
    - file_table: A pandas DataFrame object that represents the table data to be saved.
    - file_header: A dictionary with keys 'filename', 'last_mod_datetime', 'numcols', 
                    'header_rownum', 'header_guess', 'header_raw', 'header_node_text', 'plantname_in_header'
    - search_headers: A list of headers used for search indexing in the database.

    Returns:
    - A dictionary indicating the operation status for both 'table' and 'data' entries.
      Returns None if an exception occurs during database operations.
    """
    
    # Initialize a dictionary to keep track of the operation status
    status = {'table': None, 'data': None}

    try:
        # 2024-06-06 logic for multiple sheets - file_data.get('worksheet') is not None 
        tblname = file_data.get('filetoken')+'_'+file_data.get('worksheet') if file_data.get('worksheet') else file_data.get('filetoken')

        # Retrieve an existing Tables entry or None if it doesn't exist
        existing_table = Tables.query.filter_by(
            vendor=file_data.get("vendor"),
            file_name=file_data.get("filename"),
            table_name=tblname
        ).first()

        # Update existing Tables entry or create a new one
        if existing_table:
            existing_table.file_last_modified = table_header.get("last_mod_datetime")
            existing_table.dropboxurl=file_data.get("dropboxurl"),
            status['table'] = 'Updated'
        else:
            new_table = Tables(
                vendor=file_data.get("vendor"), 
                file_name=file_data.get("filename"), 
                file_last_modified = table_header.get("last_mod_datetime"),
                file_dropbox_url=file_data.get("dropboxurl"),
                table_name=tblname
            )
            db.session.add(new_table)
            status['table'] = 'New'

        # Convert the file_table DataFrame to a JSON string for storage
        df_json = file_table.to_json(orient='split')

        # Retrieve an existing TableData entry or None if it doesn't exist
        existing_table_data = TableData.query.filter_by(table_name=tblname).first()

        # Update existing TableData entry or create a new one
        if existing_table_data:
            existing_table_data.search_columns = search_headers
            existing_table_data.row_count = len(file_table)
            existing_table_data.df = df_json
            status['data'] = 'Updated'
        else:
            new_table_data = TableData(
                table_name=tblname, 
                search_columns=search_headers,
                row_count=len(file_table), 
                df=df_json
            )
            db.session.add(new_table_data)
            status['data'] = 'New'

    except SQLAlchemyError as e:
        logging.error(f"Database operation failed: {e}")
        db.session.rollback()
        return {'table': None, 'data': None}

    return status

def save_all_file_tables_in_dir(dirpath:str, use_dropbox = False):
    '''
    Top level function to create tables and save to db
    Automatically identifies unique table headings 
    '''
    try:
        load_dotenv()

        # NOTE: do NOT deploy with your key hardcoded
        tmp = os.getenv('OPENAI_API_KEY')
        os.environ["OPENAI_API_KEY"] = tmp

        # define embedding model 
        service_context = ServiceContext.from_defaults(embed_model="local")

        # get the files 
        if use_dropbox:
            yield f'Reading Dropbox files in {dirpath} at {datetime.now():%b %d %I:%M %p}:'
            dbx_refresh = os.getenv('DROPBOX_REFRESH_TOKEN')
            dbx_key = os.getenv('DROPBOX_APP_KEY')
            dbx_secret = os.getenv('DROPBOX_APP_SECRET')
            dbx_directory = os.getenv('DROPBOX_DIRECTORY')

            # logging.info(f'xx - APP_KEY: {APP_KEY}, APP_SECRET: {APP_SECRET}, REFRESH_TOKEN: {REFRESH_TOKEN}')

            if (dbx_refresh and dbx_key and dbx_secret):
                dropbox_access_token = get_dropbox_accesstoken_from_refreshtoken(dbx_refresh,dbx_key,dbx_secret)
                dbx = dropbox.Dropbox(dropbox_access_token)

                try:
                    dropbox_filenames = get_dropbox_filenames(dbx,startDir = dbx_directory)
                except Exception as err:
                    logging.error("Error occurred: %s", err)

                filenames = [f for f in dropbox_filenames.keys() if allowed_file(f.split('/')[-1])]
            else:
                raise Exception('DROPBOX_REFRESH_TOKEN, DROPBOX_APP_KEY, DROPBOX_APP_SECRET not found')
        else:
            os_filenames = get_filenames_in_directory(dirpath)
            filenames = [f for f in os_filenames.keys() if allowed_file(f.split('/')[-1])]

        # yield the status
        logging.info (f'Found {len(filenames)} files in {dirpath} at {datetime.now():%b %d %I:%M %p}:')
        yield f'Found {len(filenames)} files in {dirpath} at {datetime.now():%b %d %I:%M %p}:'
        for tmp in filenames:
            yield('  ' + str(tmp))

        # loop the files, and extract tables  
        # for full_filename in [f for n,f in enumerate(filenames) if n in [17]]:
        for full_filename in filenames: # filenames[4:6]
            # get the dirpath, filename and file type
            file_data = parse_fullfilename(full_filename = full_filename)
            # get the dropbox keys if we obtained file from dropbox 
            if use_dropbox:
                # save the dropbox objects
                file_data['dropboxdbx']=dbx
                file_data['dropboxid']=dropbox_filenames.get(full_filename)
                file_metadata = dbx.files_get_metadata(full_filename) # suggested as error check 
                shared_link = create_shared_link_with_retries(dbx, full_filename)
                file_data['dropboxurl'] = shared_link

            # check the filenames already in our dB 
            check_file = Tables.query.filter_by(
                vendor=file_data.get("vendor"),
                file_name=file_data.get("filename")
            )

            # get the first table if we found the tile
            existing_table = check_file.first() if check_file else None

            if existing_table:
                # get the elements into a list 
                file_tables = [x for x in check_file]
                # yield the status
                logging.info(f'{file_data.get("vendor")}: {file_data.get("filename")} read previously on: {existing_table.created_date:%b %d %I:%M %p}')
                yield ' '
                yield f'{file_data.get("vendor")}: {file_data.get("filename")} read previously on: {existing_table.created_date:%b %d %I:%M %p}'
                yield f'{file_data.get("filename")} last modified on {existing_table.file_last_modified} has {len(file_tables)} tables in search scope'
                continue

            # yield the status
            logging.info(f'{file_data.get("vendor")}: {file_data.get("filename")} at {datetime.now():%b %d %I:%M %p}')
            yield ' '
            yield f'{file_data.get("vendor")}: {file_data.get("filename")} at {datetime.now():%b %d %I:%M %p}'

            # branch pdf vs. xlsx files 
            # 2024-06-06 returns file_table, table_header as dicts of df, one df per worksheet 
            if file_data.get('filetype').lower() == 'pdf':
                # extract tables as dataframes - 2024-02-26 file_data has fields for dropbox  
                file_table, table_header = get_file_table_pdf(file_data)

            elif file_data.get('filetype').lower() in ['xls', 'xlsx']:
                # extract tables as dataframes - 2024-02-26 file_data has fields for dropbox 
                file_table, table_header = get_file_table_xls(file_data)
            
            # 2024-06-06 - dicts -  remove columns with all null values 
            for k,v in file_table.items():
                v.dropna(axis=1, how='all', inplace=True)
                # remove rows with all null values 
                v.dropna(axis=0, how='all', inplace=True)

                # guess the columns to search for plant names 
                search_headers = get_bestguess_table_search_columns(v)

                # sheet specific tablename if required 
                if len(file_table) > 1:
                    file_data['worksheet'] = k
                    # 2024-06-06 logic for multiple sheets - file_data.get('worksheet') is not None 
                    tblname = file_data.get('filetoken')+'_'+file_data.get('worksheet') if file_data.get('worksheet') else file_data.get('filetoken')
                else:
                    tblname = file_data.get('filetoken')

                # yield the status
                logging.info(f"Created {tblname} at {datetime.now():%b %d %I:%M %p}\nHeader: {table_header[k]['header_guess']}")
                if len(file_table) > 1: yield f"---" 
                yield f"Table: {tblname} table with {len(v)} rows "
                yield f"From: {table_header[k].get('header_raw')}"
                yield f"Guess: {table_header[k].get('header_guess')[1:-1]}"
                yield f"Search: {search_headers} "

                # save the table to the session 
                status = save_table_to_session(file_data, file_table[k], table_header[k], search_headers )

        # Commit the session to save changes to the database
        db.session.commit()

    except SQLAlchemyError as e:
        logging.error(f"save_all_file_tables_in_dir failed: {e}")
        db.session.rollback()

def background_task(app, dirpath, user_id, useDropbox=False):
    """
    Performs a background task to update file tables in a specified directory for a given user.

    Parameters:
    - app: The Flask application context.
    - dirpath: The directory path containing files to update.
    - user_id: The ID of the user for whom the updates are performed.
    - isDropbox: Boolean indicating to read files from dropbox 

    The function updates the status of file processing in the database and logs progress.
    """
    with app.app_context():
        try:
            # Check or create task completion flag for the user
            check_thread = ThreadComplete.query.get(user_id)
            if check_thread:
                check_thread.task_complete = False
            else:
                new_thread = ThreadComplete(id=user_id, task_complete=False)
                db.session.add(new_thread)
            db.session.commit()

            # Initialize or update user data
            user = User.query.get(user_id)
            user_org = Organization.query.get(user.org_id)

            if not user_org:
                logging.error(f"User org id {user.org_id} not found in Organization.")
                return
            
            # for holding details of update 
            init_log = []

            # Process each file in the directory
            for update in save_all_file_tables_in_dir(dirpath, use_dropbox=useDropbox):
                # logging.info(f"Background Update: {update}")
                init_log.append(update)
                message_queue.put(update)
            
            # Signal task completion
            completion_message = f"\nInventory Update at {datetime.now():%b %d %I:%M %p}"
            logging.info(completion_message)
            message_queue.put(completion_message)

            # Update task completion status
            if check_thread:
                check_thread.task_complete = True
                # Save the data updates
                user_org.is_init = True
                user_org.init_status = f'Last refreshed on {datetime.now():%b %d %I:%M %p}'
                user_org.init_details = "\n".join(init_log)
                db.session.merge(user_org)
                db.session.commit()
                logging.info('check_thread complete status committed to dB')

            message_queue.put(None)

        except SQLAlchemyError as e:
            db.session.rollback()
            logging.error(f"Database operation failed: {e}")
        except Exception as e:
            logging.error(f"Unexpected error: {e}")
