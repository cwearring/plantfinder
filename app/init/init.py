import base64
import io
import time 
import datetime
import pypdf  # https://pypi.org/project/pypdf/
import openpyxl as xl 
# https://medium.com/@alice.yang_10652/python-how-to-convert-excel-xls-or-xlsx-to-pdf-1b3546baba24 
import dropbox
import requests

import os
from dotenv import load_dotenv

from flask_login import current_user

from app.models import  db, UserData, User
from os import listdir
from os.path import isfile, join, getmtime

# load environment variables 
basedir = os.path.abspath(os.path.dirname(__file__))
load_dotenv(os.path.join(basedir, '.env'))

class Config:
    SECRET_KEY = os.environ.get('SECRET_KEY') or 'you-will-never-guess'
 
import requests
import os
import base64
import logging

def get_DropBox_AccessToken_from_RefreshToken():
    '''
    Get the auth token from a valid refresh token

    # https://www.dropboxforum.com/t5/Dropbox-API-Support-Feedback/Get-refresh-token-from-access-token/m-p/596755/highlight/false#M27728
    # https://www.dropbox.com/developers/documentation/http/documentation#authorization 
    # https://www.dropbox.com/oauth2/authorize?client_id=j27yo01f5f8w304&response_type=code&token_access_type=offline
    
    # i have to get the access code dynamically from the refresh token
    # https://developers.dropbox.com/oidc-guide 

    This gets a refresh token if we pass a valid AUTHORIZATIONCODEHERE
    We only need to do this once to get a persistent refresh token  
    curl https://api.dropbox.com/oauth2/token \
        -d code=AUTHORIZATIONCODEHERE \
        -d grant_type=authorization_code \
        -u APPKEYHERE:APPSECRETHERE​

    This gets the authcode from the refresh token
    curl https://api.dropbox.com/oauth2/token \
        -d refresh_token=REFRESHTOKENHERE \
        -d grant_type=refresh_token \
        -d client_id=APPKEYHERE \
        -d client_secret=APPSECRETHERE
    '''

    try:
        APP_KEY = os.environ.get('DROPBOX_APP_KEY')
        APP_SECRET = os.environ.get('DROPBOX_APP_SECRET')
        REFRESH_TOKEN = os.environ.get('DROPBOX_REFRESH_TOKEN')

        if not all([APP_KEY, APP_SECRET, REFRESH_TOKEN]):
            logging.error("Missing required Dropbox configuration.")
            return "Error: Missing configuration."

        BASIC_AUTH = base64.b64encode(f'{APP_KEY}:{APP_SECRET}'.encode()).decode()

        data = {
            "refresh_token": REFRESH_TOKEN,
            "grant_type": "refresh_token",
            "client_id": APP_KEY,
            "client_secret": APP_SECRET
        }

        response = requests.post('https://api.dropboxapi.com/oauth2/token', data=data)

        if response.status_code == 200:
            return response.json().get("access_token")
        else:
            logging.error(f"Dropbox API Error: {response.status_code} - {response.text}")
            return "Error: Failed to retrieve access token."

    except requests.RequestException as e:
        logging.error(f"Request Exception: {e}")
        return "Error: Network error occurred."

    except Exception as e:
        logging.error(f"General Exception: {e}")
        return "Error: An unexpected error occurred."

def isFileDropBox(dropboxMeta):
    return isinstance(dropboxMeta,dropbox.files.FileMetadata)

def get_onlyfilesDropBox(dbx, startDir:str, onlySubdir:list):
    filesFound = {} # key = filepath, value = dbx meta entity
    filesFoundText = '' # text blob to save printed output 

    for entry in dbx.files_list_folder(f'/{startDir}').entries:
        if isFileDropBox(entry):
            # print(f"\nIn {f'/{startDir}'} found File: {entry.name}")
            # filesFoundText += f"In {f'/{startDir}'} found File: {entry.name}\n"
            # filesFound[entry.path_lower] = entry
            junk=0
        else:
            # walk the subdirectories 
            print(f"From {f'/{startDir}'} found Directory: {entry.name}")
            # filesFoundText += f"\Found: {entry.name}\n"
            # add filter logic for shorter debug runs
            if len(onlySubdir)==0 or entry.name in onlySubdir:
                filesFoundText += f"{entry.name}:\n"
                for tst in dbx.files_list_folder(entry.path_lower).entries:
                    if isFileDropBox(tst): # it's a file 
                        print(f"{entry.name}: {tst.name}")
                        filesFoundText += f"{tst.name}\n"
                        filesFound[tst.path_lower] = tst
                        jnk = 0
                    elif False: # it's a folder - loop again
                        for tst2 in dbx.files_list_folder(tst.path_lower).entries:
                            if isFileDropBox(tst2):
                                print(f"{entry.name}: {tst2.name}")
                                filesFoundText += f"{tst2.name}\n"
                                filesFound[tst2.path_lower] = tst2
                            else:
                                print(f"In {tst.name} found Folder: {tst2.name}")
                                # print(f"Files from Folder: {tst2.name} are not loaded")
                                # filesFoundText += f"In {tst.name} found Folder: {tst2.name}\n"
                                # filesFoundText += f"Files from Folder: {tst2.name} are not loaded\n"
                            jnk = 0

    return(filesFound, filesFoundText)

def createTdictDropBox(dbx, startDir:str, onlySubdir:list):
    '''
    Input: Token for the dropbox root  
    Output: Dictionary with key as file types and value as list of dropbox file metadata
    '''
    # get files in directory 
    tmpDbxFiles, tmpDbxFilesText = get_onlyfilesDropBox(dbx, startDir, onlySubdir)

   # create a dict with key as file types and value as list of dropbox file objects
    key = set([x.split('.')[-1] for x in tmpDbxFiles.keys()])
    tmp = [(k.split('.')[-1], v) for k,v in tmpDbxFiles.items()]
    tdict = {k:[v[1] for v in tmp if v[0] == k ] for k in key}

    return(tdict, tmpDbxFilesText)

def getDFdictDropBox(tdict:dict, dbx):
    '''
    Input: tdict - dictionary wity key = file type and value = dropbox file object 
    Output: dfDict - dictionary key = filePath, value = text blog of contents 
    '''
    dfDict = {}
    dfDictFiles = []
    # Read each pdf into list of DataFrame
    for file_type in tdict.keys():
        if 'pdf' in file_type:
            print(f"Processing {len(tdict[file_type])} {file_type} files")
            for mfile in tdict[file_type]:
                # dfDictFiles.append((f"{mfile.path_lower.split('/')[-2]}: {mfile.name} \n"))
                dfDictFiles.append((f"{mfile.name} \n"))
                try:
                    md,res = dbx.files_download(mfile.path_lower)
                    myPdfMod = mfile.client_modified
                    with io.BytesIO(res.content) as open_pdf_file:
                        reader = pypdf.PdfReader(open_pdf_file)
                        dfs = ""
                        for n in range(len(reader.pages)):
                            page = reader.pages[n]
                            dfs += page.extract_text()
                    #  extract all text to variable
                except:
                    dfs = 'Python Error'

                dfDict[mfile.path_lower] = [dfs, myPdfMod]

        elif 'xls' in file_type:
            print(f"Processing {len(tdict[file_type])} {file_type} files")
            for mfile in tdict[file_type]:
                # dfDictFiles.append((f"{mfile.path_lower.split('/')[-2]}: {mfile.name} \n"))
                dfDictFiles.append((f"{mfile.name} \n"))
                try:
                    md,res = dbx.files_download(mfile.path_lower)
                    xl_bytes =  io.BytesIO(res.content)
                    myXl = xl.load_workbook(xl_bytes, data_only=True)
                    myXLmod = [myXl.properties.modified.strftime("%B %d, %Y"), myXl.properties.lastModifiedBy]                     
                    dfs = ""
                    for myTab in myXl._sheets:
                        # dfs = (f"{mfile} sheet {myTab} is {myTab.max_column} columns by {myTab.max_row} rows", myTab)
                        for row in myTab.iter_rows(min_row=1, values_only=False):
                            xxx = [str(x._value) for x in row if (x._value and x._value != 'None')]
                            if len(xxx) >0:
                                dfs += f"{' '.join(xxx)}\n"
                        jnk = 0 
                except:
                    dfs = f'Python Error'
                               
                dfDict[mfile.path_lower] = [dfs, myXLmod]

    return(dfDict, dfDictFiles)

    '''
    Input: Filepath from current directory 
    Output: timestamp for last modified date-time
    '''
    # ti_c = os.path.getctime(f"{mypath}/{myTestFile}")  # c_ti = time.ctime(ti_c)
    ti_m = getmtime(f"{myFilePath}")
    # Converting the time in seconds to a timestamp
    m_ti = time.ctime(ti_m)
    return (m_ti)

    '''
    Input: Path to directory containing both PDF and xlsx files 
    Output: Dictionary with key as file types and value as list of filePaths
    '''
    # get files in directory 
    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]

   # create a dict with key as file types and value as list of files 
    key = set([x.split('.')[1] for x in onlyfiles])
    tmp = [(x.split('.')[1], x) for x in onlyfiles]
    tdict = {k:[join(mypath,v[1]) for v in tmp if v[0] == k ] for k in key}
    return(tdict)

def initSearchDropBox(onlySubdir:list):

    # this code creates (dbx,startDir) that were passed as arguments 
    MY_ACCESS_TOKEN = get_DropBox_AccessToken_from_RefreshToken()
    dbx = dropbox.Dropbox(MY_ACCESS_TOKEN)
    startDir = 'Garden Centre Ordering Forms'

    # Get the user_data_id
    user_id = current_user.id
 
    # init and save the dfDict for the dropbox root directory mypath
    if dbx:
        tdict, tdictText = createTdictDropBox(dbx, startDir, onlySubdir)
        dfDict, dfDictFileText = getDFdictDropBox(tdict=tdict, dbx=dbx)

        # define the return payload
        now = datetime.datetime.now()
        formatted_date_time = f"{now.strftime('%I:%M %p')} on {now.strftime('%A %b%d %Y')}"
        initStatus = f"Inventory refreshed at {formatted_date_time}"

        # save in the database with key = user_id
        user_data = UserData(id=user_id, data={ 'status':initStatus,
                                                'tdict': tdict,
                                                'dfDict': dfDict})
        db.session.merge(user_data)
        db.session.commit()

        initComplete = {
            'status': initStatus,
            'filesfound': tdictText,
            'filesloaded': dfDictFileText
        }
    else:

        initComplete = "Need a dropbox token and root directory for suppliers' files"

    return (initComplete)

def XXXget_DropBox_AccessToken_from_RefreshToken():

    APP_KEY = os.environ.get('DROPBOX_APP_KEY') or 'nodropboxkey'
    APP_SECRET = os.environ.get('DROPBOX_APP_SECRET') or 'nodropboxsecret'
    BASIC_AUTH = base64.b64encode(f'{APP_KEY}:{APP_SECRET}'.encode())

    # using the refresh token for now 
    REFRESH_TOKEN = os.environ.get('DROPBOX_REFRESH_TOKEN') or 'nodropboxrefreshtoken'

    
    data = {
            "refresh_token":REFRESH_TOKEN,
            "grant_type":"refresh_token",
            "client_id":APP_KEY,
            "client_secret":APP_SECRET
    }

    # Use the refresh token to get a valid temp auth token
    response = requests.post('https://api.dropboxapi.com/oauth2/token',
                            data=data)
    
    if response.status_code == 200:
        return(response.json()["access_token"])
    else:
        return(f"{response.status_code} error {response.text}")
        #print(json.dumps(json.loads(response.text), indent=2))

