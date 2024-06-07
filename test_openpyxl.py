import openpyxl
import pandas as pd 
from openpyxl import load_workbook

def read_all_data(path):
    workbook = load_workbook(filename=path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"Title = {sheet.title}")
        for value in sheet.iter_rows(values_only=True):
            print(value)

if __name__ == "__main__":
    # x = pd.read_excel("/Users/cwearring/Downloads/NVK_trees_May25.xls")
    read_all_data("/Users/cwearring/Downloads/NVK+availability+May+25.xlsx")